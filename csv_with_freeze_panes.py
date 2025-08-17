import csv
import os
import sys
import argparse

def create_excel_friendly_csv_with_freeze(csv_path, output_path=None):
    """ExcelでヘッダーがフリーズするCSVファイルを作成"""
    
    if not os.path.exists(csv_path):
        print(f"❌ CSVファイルが見つかりません: {csv_path}")
        return
    
    # 出力パスを決定
    if output_path is None:
        base_name = os.path.splitext(csv_path)[0]
        output_path = f"{base_name}_freeze.csv"
    
    # 元のCSVを読み込み
    with open(csv_path, 'r', encoding='utf-8-sig') as f:
        content = f.read()
    
    # UTF-8 BOMでCSVを保存（Excelで正しく開くため）
    with open(output_path, 'w', encoding='utf-8-sig', newline='') as f:
        f.write(content)
    
    # 同じディレクトリにExcel用のマクロファイル（.xlsm）も作成
    xlsm_path = output_path.replace('.csv', '_with_freeze.xlsm')
    create_excel_with_freeze_panes(csv_path, xlsm_path)
    
    print(f"✅ Excel対応CSV作成: {output_path}")
    print(f"✅ ヘッダー固定Excel作成: {xlsm_path}")
    print("📋 Excelでヘッダーを固定するには:")
    print("   1. Excelでファイルを開く")
    print("   2. A2セルを選択")
    print("   3. [表示]タブ → [ウィンドウ枠の固定] → [ウィンドウ枠の固定]")
    
    return output_path, xlsm_path

def create_excel_with_freeze_panes(csv_path, xlsm_path):
    """XLSXファイルでヘッダー固定を実装"""
    try:
        import openpyxl
        from openpyxl.utils.dataframe import dataframe_to_rows
        import pandas as pd
        
        # CSVをDataFrameとして読み込み
        df = pd.read_csv(csv_path, encoding='utf-8-sig')
        
        # 新しいワークブックを作成
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"
        
        # DataFrameをワークシートに書き込み
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # ヘッダー行をフリーズ（A2から下をスクロール可能に）
        ws.freeze_panes = 'A2'
        
        # ヘッダーのスタイルを設定
        from openpyxl.styles import Font, PatternFill
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for cell in ws[1]:  # 1行目（ヘッダー）
            cell.font = header_font
            cell.fill = header_fill
        
        # 列幅を自動調整
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # 最大50文字
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # ファイルを保存
        wb.save(xlsm_path.replace('.xlsm', '.xlsx'))
        print(f"✅ Excel形式（ヘッダー固定）: {xlsm_path.replace('.xlsm', '.xlsx')}")
        
    except ImportError:
        print("⚠️ openpyxl, pandasがインストールされていません")
        print("   pip install openpyxl pandas でインストールしてください")
        create_simple_excel_instructions(csv_path, xlsm_path)

def create_simple_excel_instructions(csv_path, xlsm_path):
    """シンプルな手順書を作成"""
    instructions_path = xlsm_path.replace('.xlsm', '_instructions.txt')
    
    with open(instructions_path, 'w', encoding='utf-8') as f:
        f.write("""Excel でヘッダーを固定する手順:

1. Excelでファイルを開く
2. データの2行目（A2セル）をクリック
3. [表示]タブをクリック
4. [ウィンドウ枠の固定]をクリック
5. [ウィンドウ枠の固定]を選択

これで1行目（ヘッダー）が固定され、
下にスクロールしてもヘッダーが表示されます。

または:
- [データ]タブ → [フィルター]で▼ボタン付きヘッダーにする
- [挿入]タブ → [テーブル]でテーブル形式にする
""")
    
    print(f"📋 手順書作成: {instructions_path}")

def main():
    parser = argparse.ArgumentParser(description='ExcelでヘッダーがフリーズするCSVを作成')
    parser.add_argument('csv_file', help='変換するCSVファイル')
    parser.add_argument('-o', '--output', help='出力CSVファイル名')
    args = parser.parse_args()
    
    create_excel_friendly_csv_with_freeze(args.csv_file, args.output)

if __name__ == "__main__":
    if len(sys.argv) == 1:
        # 引数なしの場合、例として26147のCSVを変換
        csv_file = r"C:\Users\bnr39\OneDrive\カルテOCR\26147\vision_iop_26147.csv"
        if os.path.exists(csv_file):
            create_excel_friendly_csv_with_freeze(csv_file)
        else:
            print("使用法: python csv_with_freeze_panes.py <CSVファイル> [-o 出力ファイル]")
    else:
        main()
